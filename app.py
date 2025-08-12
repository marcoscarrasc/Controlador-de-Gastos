from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import json

app = Flask(__name__)

# Configuración
EXCEL_FILE = 'gastos.xlsx'
DATA_FOLDER = 'data'

# Crear carpeta de datos si no existe
if not os.path.exists(DATA_FOLDER):
    os.makedirs(DATA_FOLDER)

EXCEL_PATH = os.path.join(DATA_FOLDER, EXCEL_FILE)

def load_expenses():
    """Cargar gastos desde el archivo Excel"""
    if os.path.exists(EXCEL_PATH):
        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            
            expenses = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:  # Verificar que no sea una fila vacía
                    expense = {
                        'id': int(row[0]),
                        'fecha': row[1],
                        'descripcion': row[2],
                        'categoria': row[3],
                        'monto': float(row[4]),
                        'mes': int(row[5]),
                        'año': int(row[6])
                    }
                    expenses.append(expense)
            return expenses
        except Exception as e:
            print(f"Error loading expenses: {e}")
            return []
    return []

def save_expenses(expenses):
    """Guardar gastos en el archivo Excel"""
    wb = Workbook()
    ws = wb.active
    
    # Encabezados
    headers = ['id', 'fecha', 'descripcion', 'categoria', 'monto', 'mes', 'año']
    ws.append(headers)
    
    # Datos
    for expense in expenses:
        row = [
            expense['id'],
            expense['fecha'],
            expense['descripcion'],
            expense['categoria'],
            expense['monto'],
            expense['mes'],
            expense['año']
        ]
        ws.append(row)
    
    wb.save(EXCEL_PATH)

def get_next_id():
    """Obtener el siguiente ID disponible"""
    expenses = load_expenses()
    if not expenses:
        return 1
    return max(expense['id'] for expense in expenses) + 1

@app.route('/')
def index():
    """Página principal"""
    return render_template('index.html')

@app.route('/api/expenses', methods=['GET'])
def get_expenses():
    """Obtener todos los gastos"""
    expenses = load_expenses()
    return jsonify(expenses)

@app.route('/api/expenses', methods=['POST'])
def add_expense():
    """Agregar un nuevo gasto"""
    try:
        data = request.json
        expenses = load_expenses()
        
        # Crear nuevo gasto
        new_expense = {
            'id': get_next_id(),
            'fecha': data['fecha'],
            'descripcion': data['descripcion'],
            'categoria': data['categoria'],
            'monto': float(data['monto']),
            'mes': datetime.strptime(data['fecha'], '%Y-%m-%d').month,
            'año': datetime.strptime(data['fecha'], '%Y-%m-%d').year
        }
        
        expenses.append(new_expense)
        save_expenses(expenses)
        
        return jsonify({'success': True, 'expense': new_expense})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/expenses/<int:expense_id>', methods=['PUT'])
def update_expense(expense_id):
    """Actualizar un gasto existente"""
    try:
        data = request.json
        expenses = load_expenses()
        
        # Buscar y actualizar el gasto
        for expense in expenses:
            if expense['id'] == expense_id:
                expense.update({
                    'fecha': data['fecha'],
                    'descripcion': data['descripcion'],
                    'categoria': data['categoria'],
                    'monto': float(data['monto']),
                    'mes': datetime.strptime(data['fecha'], '%Y-%m-%d').month,
                    'año': datetime.strptime(data['fecha'], '%Y-%m-%d').year
                })
                save_expenses(expenses)
                return jsonify({'success': True, 'expense': expense})
        
        return jsonify({'success': False, 'error': 'Gasto no encontrado'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/expenses/<int:expense_id>', methods=['DELETE'])
def delete_expense(expense_id):
    """Eliminar un gasto"""
    try:
        expenses = load_expenses()
        expenses = [expense for expense in expenses if expense['id'] != expense_id]
        save_expenses(expenses)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/expenses/monthly/<int:year>/<int:month>', methods=['GET'])
def get_monthly_expenses(year, month):
    """Obtener gastos de un mes específico"""
    expenses = load_expenses()
    monthly_expenses = [
        expense for expense in expenses 
        if expense['año'] == year and expense['mes'] == month
    ]
    return jsonify(monthly_expenses)

@app.route('/api/expenses/summary', methods=['GET'])
def get_expenses_summary():
    """Obtener resumen de gastos por mes"""
    expenses = load_expenses()
    if not expenses:
        return jsonify([])
    
    # Agrupar por mes y año
    summary = {}
    for expense in expenses:
        key = f"{expense['año']}-{expense['mes']:02d}"
        if key not in summary:
            summary[key] = {
                'año': expense['año'],
                'mes': expense['mes'],
                'total': 0,
                'cantidad': 0
            }
        summary[key]['total'] += expense['monto']
        summary[key]['cantidad'] += 1
    
    return jsonify(list(summary.values()))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
